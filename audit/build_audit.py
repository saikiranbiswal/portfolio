#!/usr/bin/env python3
"""Build the portfolio audit workbook from a grounded issue list.

Every issue below was verified against the live site (rendered in a browser
preview) and/or the source tree on 2026-06-24. Re-run to regenerate the xlsx.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date

# ---------------------------------------------------------------- palette
INK      = "1A1A2E"   # near-black headers
CORAL    = "E8825A"   # site accent
SLATE    = "2B2B40"
LIGHT    = "F4F1EC"
P0_FILL  = "F8CBAD"   # red-ish
P1_FILL  = "FFE699"   # amber
P2_FILL  = "C6E0B4"   # green
HDR_FILL = "1A1A2E"
ZEBRA    = "FAF8F5"

thin = Side(style="thin", color="D9D2C7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfont(sz=11, color="FFFFFF", bold=True):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

# ---------------------------------------------------------------- data
# Columns: ID, Title, Category, Page/Location, Type, Severity, Impact, Effort,
#          Priority, Evidence, Recommended Fix, Status, Owner
COLS = ["ID","Title","Category","Page / Location","Type","Severity",
        "Impact","Effort","Priority Score","Evidence (what was observed)",
        "Recommended Fix","Status","Owner"]

# Impact 1-5, Effort 1-5 (1=small). Priority = round(Impact / Effort * 10 ... )
ISSUES = [
 ["P-01","Mobile nav overflows — 'Contact' is cut off, no hamburger menu","Navigation & IA",
  "All pages (global header) @ ≤420px","Bug","P0",5,2,
  "On a 375px viewport the horizontal nav runs off-screen; the 'Contact' item is clipped to 'C…' and there is no responsive/collapse menu. Reachability of a key CTA fails on phones.",
  "Add a responsive breakpoint: collapse nav into a hamburger/drawer under ~640px; ensure all six items + status remain reachable.","Open","Sai"],

 ["P-02","Homepage value prop & headline live inside <canvas> — invisible to SEO & screen readers","SEO & Accessibility",
  "index.html (hero gallery)","Bug","P0",5,4,
  "The entire hero (project cards + the 'Products built from problem to proof.' headline) is rendered on a <canvas>. The DOM exposes only a 'Skip the interactive gallery' link, nav, a caption and the theme toggle. No crawlable headline, no real anchors.",
  "Render the headline and a text fallback list of projects as real HTML behind/below the canvas (progressive enhancement). Keep canvas as a visual layer only.","Open","Sai"],

 ["P-03","Inconsistent headline stats across pages (8 vs 9 products, 2 vs 3 case studies)","Content & Consistency",
  "products.html vs labs.html","Content","P0",4,1,
  "Work page states '3 CASE STUDIES / 8 WORKING PRODUCTS / 7 YEARS'. Labs page states '5 LABS / 9 PRODUCTS / 7 YEARS' and 'two flagship case studies'. The product and case-study counts do not reconcile.",
  "Decide the canonical numbers and propagate from a single source (content/site.json). A VP will notice mismatched self-reported metrics.","Open","Sai"],

 ["P-04","Job title is inconsistent across the site & résumé","Content & Consistency",
  "Global / case studies / résumé","Content","P0",4,2,
  "'Enterprise · AI Product Manager' (nav/meta), 'Product Lead' (Collections case study Role field), 'Product Consultant' (résumé PDF title, per git history), 'lead PM roles' (status pill). Four different self-descriptions.",
  "Choose ONE primary title and apply everywhere incl. the résumé PDF and case-study Role fields.","Open","Sai"],

 ["P-05","No og:image / Twitter card on any page — links show no preview when shared","SEO & Accessibility",
  "All top-level pages","SEO","P0",4,2,
  "grep shows 0 og:image and 0 twitter:* tags on index/products/artifacts/labs/about/contact. When the link is pasted into LinkedIn, Slack, or iMessage it renders with no preview image — the exact channels a recruiter uses.",
  "Add og:image (1200×630), og:url, twitter:card=summary_large_image to every page. Generate a branded share card.","Open","Sai"],

 ["P-06","Mobile homepage gallery is unreadable — cards overlap, hero crowded","Mobile & Responsive",
  "index.html @ 375px","UX","P0",4,3,
  "At 375px the floating project cards overlap heavily, micro-text is illegible, and the headline collides with the cards. First impression on phones is visual noise rather than a clear value prop.",
  "Provide a mobile-specific layout: replace the free-floating canvas with a clean stacked list of projects under the headline.","Open","Sai"],

 ["P-07","IA redundancy: 'Work', 'Labs', and 'Artifacts' overlap and confuse the taxonomy","Navigation & IA",
  "Global nav","UX","P1",4,3,
  "Work (products.html) and Labs (labs.html) both present the same flagship case studies and an overlapping product list; Artifacts is a third 'evidence' bucket. Three sections covering similar material dilutes the story.",
  "Collapse to a clearer model, e.g. Work (case studies) / Playground (interactive apps) / Artifacts (downloads), with no content shown in two places.","Open","Sai"],

 ["P-08","Canvas project cards are not real links / not keyboard-navigable","SEO & Accessibility",
  "index.html hero","Bug","P1",3,3,
  "Projects are drawn on canvas; there are no per-project <a> elements, so keyboard and screen-reader users cannot tab to or open individual projects (only the global 'skip' link works).",
  "Mirror each project as a focusable, labelled link in the DOM fallback added in P-02.","Open","Sai"],

 ["P-09","Case-study status ambiguity — FASTag is a 3rd case study but demoted to 'Selected Work'","Content & Consistency",
  "products.html","Content","P1",3,2,
  "Two 'Flagship Case Studies' (Collections, LOS) are featured, while FASTag — also a full case study with its own page — appears only in the 'Selected Work' list as item 01. Inconsistent with the '3 case studies' stat.",
  "Either promote FASTag to flagship or relabel the stat; make case-study tiering explicit and consistent.","Open","Sai"],

 ["P-10","labs.html metadata contradicts itself (five vs six labs)","Content & Consistency",
  "labs.html <head> + body","Content","P1",2,1,
  "<title> says 'Five discipline labs', og:title says 'Two flagships, six labs', and the on-page stat says '5 LABS'. Three different numbers in one page.",
  "Pick the correct lab count and align title, og:title, and the visible stat.","Open","Sai"],

 ["P-11","Brand accent color inconsistent — case studies use blue, main site uses coral","Visual & Brand",
  "case-studies/*.html vs main site","UX","P1",3,2,
  "The case-study template (e.g. collections-cloud) uses a blue accent and a distinct header chrome, while the rest of the site uses a coral/orange accent. Feels like two different products.",
  "Unify the accent token and header treatment across case-study pages and the main shell.","Open","Sai"],

 ["P-12","Duplicate / orphaned artifact directories bloat the repo","Technical & Hygiene",
  "artifacts/ vs assets/artifacts/files/","Tech","P1",2,2,
  "The Artifacts page serves files from assets/artifacts/files/ (all 9 verified present), but a separate artifacts/ tree (decks, docx, xlsx duplicated under multiple subfolders) is also committed and appears unused.",
  "Confirm the canonical location, delete the orphaned tree, keep one source of truth.","Open","Sai"],

 ["P-13","CMS server files committed / present in deploy surface","Technical & Hygiene",
  "cms_server.py, .cms_server.log, admin.html","Tech","P1",3,2,
  "cms_server.py and a runtime .cms_server.log sit in the repo root; admin.html (CMS) is published (noindex only). Server-side script + logs should not ship with a static site, and the log was re-introduced after an earlier .gitignore was reverted in the rebase reset.",
  "Re-add .cms_server.log to .gitignore, exclude cms_server.py from the deployed artifact, and confirm admin.html is not linked anywhere.","Open","Sai"],

 ["P-14","Public product claims not reconciled to actual stage / evidence","Content & Consistency",
  "products.html / case studies","Content","P1",3,3,
  "Self-flagged in the Atlas board (ATL-201 'Reconcile portfolio product claims'). Some products read as shipped while case-study Stage fields say 'Prototype · In testing'. Mixed signals on what is live vs concept.",
  "Add an explicit, consistent stage label (Concept / Prototype / In testing / Live) to every product and case study.","Open","Sai"],

 ["P-15","Artifacts page wastes ~340px of empty space above the hero (desktop)","Visual & Brand",
  "artifacts.html","UX","P1",2,2,
  "On desktop the hero copy ('The work behind the work.') starts ~340px down, leaving a large empty dark band above the fold and pushing content below it.",
  "Tighten the hero top spacing / vertical rhythm so the headline sits in the upper third.","Open","Sai"],

 ["P-16","No clear primary CTA / next step for a hiring manager on the homepage","Navigation & IA",
  "index.html","UX","P1",4,3,
  "The home experience is a novel 'drag to explore' canvas with no obvious primary action (View work / Read a case study / Contact). Hiring managers skim — an unclear next step risks a bounce before the work is seen.",
  "Add a prominent primary CTA (e.g. 'See the work' → case studies) alongside the interactive gallery.","Open","Sai"],

 ["P-17","Nav label 'Work' points to products.html (label/URL mismatch)","Navigation & IA",
  "Global nav","Tech","P2",1,1,
  "The 'Work' nav item links to products.html. Minor, but the URL slug doesn't match the label and 'products' under-sells case-study content.",
  "Rename the file to work.html (with a redirect) or accept and document; low priority.","Open","Sai"],

 ["P-18","Legacy / orphaned pages still present (gallery.html, project.html)","Technical & Hygiene",
  "gallery.html, project.html","Tech","P2",2,2,
  "gallery.html and project.html exist and load but project.html is not in the primary nav; unclear if they are still intended. Dead pages dilute and can be indexed.",
  "Remove if retired, or wire them in intentionally; add noindex if kept as scratch.","Open","Sai"],

 ["P-19","Homepage cards bleed off-screen with unreadable micro-text; one card is a recursive screenshot of the portfolio","Visual & Brand",
  "index.html hero","UX","P2",2,2,
  "Several floating cards are clipped at the viewport edges showing illegible micro-text, and one card is a screenshot of the portfolio homepage itself (portfolio-inside-portfolio), which reads as filler.",
  "Curate the card set: drop the recursive card, ensure each visible card is legible or clearly decorative.","Open","Sai"],

 ["P-20","Résumé PDF title says 'Product Consultant' — align with chosen site title","Content & Consistency",
  "assets/resume.pdf","Content","P2",2,1,
  "Per git history the résumé file metadata/title is 'Sai Kiran Biswal — Product Consultant', conflicting with the site's 'Product Manager' positioning (see P-04).",
  "Regenerate the résumé with the single chosen title and consistent headline.","Open","Sai"],

 ["P-21","Atlas 'source gap' — confirm Atlas is a real prototype, not a placeholder","Technical & Hygiene",
  "apps/atlas/","Tech","P2",2,2,
  "Self-flagged in the Atlas board (ATL-196 'Resolve Atlas source gap' / 'Publish a real preview and label the system as a prototype'). Ensure the linked app is substantive and labelled.",
  "Publish a real preview/state and a clear 'prototype' label, or remove from the featured list.","Open","Sai"],

 ["P-22","Verify outbound LinkedIn URL resolves","Content & Consistency",
  "contact.html","Content","P2",2,1,
  "Contact links to https://www.linkedin.com/in/sai-kiran-biswal (opens in new tab). Confirm the vanity URL is correct and live — a broken primary social link costs credibility.",
  "Manually verify the LinkedIn handle resolves; fix if the slug differs.","Open","Sai"],

 ["P-23","'Open to lead PM roles' status must match the chosen title","Content & Consistency",
  "Global header status pill","Content","P2",1,1,
  "The status pill says 'Open to lead PM roles' — fine, but ensure it agrees with whatever single title is chosen in P-04 (PM vs Lead vs Consultant).",
  "Align wording once the canonical title is set.","Open","Sai"],
]

# priority score helper
def pscore(impact, effort):
    return round(impact / effort * 2, 1)  # higher = do first

# ---------------------------------------------------------------- workbook
wb = openpyxl.Workbook()

def style_header(ws, ncols, row=1):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=HDR_FILL)
        cell.font = hfont(11, "FFFFFF", True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 28

def sev_fill(sev):
    return {"P0":P0_FILL,"P1":P1_FILL,"P2":P2_FILL}.get(sev, "FFFFFF")

def write_table(ws, rows, title=None):
    start = 1
    ws.append(COLS)
    style_header(ws, len(COLS))
    for r in rows:
        impact, effort = r[6], r[7]
        row = r[:8] + [pscore(impact, effort)] + r[9:]
        # insert priority at index 8 (already accounted): rebuild properly
        ws.append(r[0:8] + [pscore(impact, effort)] + r[9:])
    # widths
    widths = [7,34,20,26,10,9,8,8,9,52,46,9,8]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # style body
    for ri in range(2, ws.max_row+1):
        sev = ws.cell(row=ri, column=6).value
        for ci in range(1, len(COLS)+1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ci == 6:
                cell.fill = PatternFill("solid", fgColor=sev_fill(sev))
                cell.font = Font(bold=True)
            elif ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        ws.row_dimensions[ri].height = 70
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

# ---- Master sheet
master = wb.active
master.title = "Master Issue Log"
write_table(master, ISSUES)

# ---- Category sheets
cats = {}
for r in ISSUES:
    cats.setdefault(r[2], []).append(r)
SHEET_NAMES = {
 "Navigation & IA":"Nav & IA",
 "Content & Consistency":"Content",
 "SEO & Accessibility":"SEO & A11y",
 "Mobile & Responsive":"Mobile",
 "Visual & Brand":"Visual & Brand",
 "Technical & Hygiene":"Tech & Hygiene",
}
for cat, rows in cats.items():
    ws = wb.create_sheet(SHEET_NAMES.get(cat, cat[:28]))
    write_table(ws, rows)

# ---- Dashboard
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False

dash["B2"] = "Portfolio Audit — saikiranbiswal.github.io/portfolio"
dash["B2"].font = Font(size=18, bold=True, color=INK)
dash["B3"] = f"Prepared {date.today().isoformat()}  ·  Goal: make a PM / hiring manager / VP of Product hit an 'aha' moment"
dash["B3"].font = Font(size=11, italic=True, color="666666")

# headline counters
total = len(ISSUES)
by_sev = {"P0":0,"P1":0,"P2":0}
for r in ISSUES: by_sev[r[5]] += 1
by_cat = {}
for r in ISSUES: by_cat[r[2]] = by_cat.get(r[2],0)+1

cards = [("Total open issues", total, CORAL),
         ("P0 — Critical", by_sev["P0"], "C0392B"),
         ("P1 — Important", by_sev["P1"], "B7950B"),
         ("P2 — Polish", by_sev["P2"], "1E8449")]
col = 2
for label, val, color in cards:
    c = dash.cell(row=5, column=col); c.value = label
    c.font = Font(size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center")
    v = dash.cell(row=6, column=col); v.value = val
    v.font = Font(size=26, bold=True, color=color)
    v.alignment = Alignment(horizontal="center")
    dash.column_dimensions[get_column_letter(col)].width = 20
    col += 1

# severity table for chart
dash["B9"] = "By Severity"; dash["B9"].font = Font(bold=True, size=12, color=INK)
dash["B10"]="Severity"; dash["C10"]="Count"
for i,(k) in enumerate(["P0","P1","P2"]):
    dash.cell(row=11+i, column=2, value=k)
    dash.cell(row=11+i, column=3, value=by_sev[k])

# category table for chart
dash["E9"] = "By Category"; dash["E9"].font = Font(bold=True, size=12, color=INK)
dash["E10"]="Category"; dash["F10"]="Count"
ci=0
for k,v in sorted(by_cat.items(), key=lambda x:-x[1]):
    dash.cell(row=11+ci, column=5, value=k)
    dash.cell(row=11+ci, column=6, value=v)
    ci+=1
dash.column_dimensions["E"].width = 24

# pie chart severity
pie = PieChart(); pie.title = "Issues by Severity"
data = Reference(dash, min_col=3, min_row=10, max_row=13)
labels = Reference(dash, min_col=2, min_row=11, max_row=13)
pie.add_data(data, titles_from_data=True); pie.set_categories(labels)
pie.height=7; pie.width=9
dash.add_chart(pie, "B16")

# bar chart category
bar = BarChart(); bar.type="bar"; bar.title="Issues by Category"
bdata = Reference(dash, min_col=6, min_row=10, max_row=10+ci)
blab = Reference(dash, min_col=5, min_row=11, max_row=10+ci)
bar.add_data(bdata, titles_from_data=True); bar.set_categories(blab)
bar.height=7; bar.width=12; bar.legend=None
dash.add_chart(bar, "E16")

# top priorities list
dash["B32"]="Top priorities (start here)"; dash["B32"].font=Font(bold=True,size=12,color=INK)
top = sorted(ISSUES, key=lambda r: (r[5], -pscore(r[6],r[7])))
trow=33
dash.cell(row=trow,column=2,value="ID").font=hfont(10,"FFFFFF")
dash.cell(row=trow,column=3,value="Issue").font=hfont(10,"FFFFFF")
dash.cell(row=trow,column=6,value="Sev").font=hfont(10,"FFFFFF")
for c in (2,3,6):
    dash.cell(row=trow,column=c).fill=PatternFill("solid",fgColor=HDR_FILL)
for r in [x for x in top if x[5]=="P0"]:
    trow+=1
    dash.cell(row=trow,column=2,value=r[0])
    dash.cell(row=trow,column=3,value=r[1])
    sc=dash.cell(row=trow,column=6,value=r[5])
    sc.fill=PatternFill("solid",fgColor=sev_fill(r[5]))
    sc.font=Font(bold=True)
dash.merge_cells(start_row=34,start_column=3,end_row=trow,end_column=5) if False else None

# README / methodology
rd = wb.create_sheet("Methodology")
rd.sheet_view.showGridLines=False
rd.column_dimensions["B"].width=110
notes = [
 ("How this audit was produced", True),
 ("", False),
 ("The live site was rendered in a browser preview and every top-level page, the case-study", False),
 ("template, and the Atlas app were inspected on desktop (1280px) and mobile (375px). The", False),
 ("source tree was checked for broken links, orphaned files, SEO/meta tags, and file integrity.", False),
 ("", False),
 ("Severity key", True),
 ("P0 — Critical: breaks credibility or reachability; fix before sharing the link with a VP.", False),
 ("P1 — Important: noticeably weakens the story or hygiene; fix this cycle.", False),
 ("P2 — Polish: refinement; fix when time allows.", False),
 ("", False),
 ("Priority Score = Impact / Effort × 2 (higher = do first). Impact & Effort scored 1–5 (1 = small effort).", False),
 ("", False),
 ("Note on overlaps", True),
 ("P-04 (title), P-20 (résumé title) and P-23 (status pill) are one decision — pick a single title once.", False),
 ("P-02 (canvas SEO) and P-08 (canvas a11y) share the same fix (a DOM fallback).", False),
 ("Several items (P-14 reconcile claims, P-21 Atlas gap) were already self-identified on the Atlas board.", False),
]
for i,(t,bold) in enumerate(notes, start=2):
    c=rd.cell(row=i,column=2,value=t)
    c.font=Font(bold=bold, size=13 if bold else 11, color=INK if bold else "333333")

# order sheets: Dashboard, Master, categories, methodology
wb.move_sheet("Dashboard", -wb.sheetnames.index("Dashboard"))

out = "audit/portfolio-audit.xlsx"
wb.save(out)
print("Wrote", out, "with sheets:", wb.sheetnames)
print("Total issues:", total, "| P0:",by_sev["P0"],"P1:",by_sev["P1"],"P2:",by_sev["P2"])
